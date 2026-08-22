#!/usr/bin/env python3
"""Extrae `knowledge/roc_ici.yaml`: el %ROC (Return of Capital) DEFINITIVO por ticker y año
fiscal, desde los reportes "ICI Primary Layout" que YieldMax publica en
https://yieldmaxetfs.com/tax-documents/ (públicos, sin login).

Por qué existe además de `fetch_roc_19a.py`: los avisos 19(a) que ese script trae son la
ESTIMACIÓN del gestor al momento de pagar. El ICI es el carácter fiscal DEFINITIVO tras el
cierre del año (lo que de verdad determina la casilla 3 del 1099-DIV / la reclasificación del
1042-S). Ver `Obsidian/IA/traspaso-2026-08-21-roc-historico-ici.md` para el porqué completo y
el ground truth de 2024 verificado a mano.

Misma forma que `fetch_roc_19a.py`: descarga separada del parseo, para poder testear el
parser sin red (ver `test_roc_ici.py`, con fixtures reales/recortadas en `fixtures/`).

⛔ Esta tarea SOLO produce `knowledge/roc_ici.yaml` y datos para el reporte de contraste.
No cablea nada a `ui/adapters.py` ni a `_roc_pct_by_year`, y no toca `knowledge/roc_19a.yaml`.

El layout de cada documento (medido, no supuesto — ver el traspaso):
  - 2024 (.xlsx): hoja "Primary Layout FINAL", cabecera fila 13, datos desde fila 17.
  - 2025 (.xlsx): hoja "YM - ICI Primary Layout" (+ hoja "Cover" sin datos), cabecera fila 12,
    datos desde fila 15.
  - En AMBOS años las columnas que importan caen en las MISMAS posiciones (verificado):
    C=Ticker Symbol, H=Ex-Dividend Date, J=Total Distribution Per Share, K/L/M=atribución de
    año (Prior/Next/Current — el año que declara cada cabecera cambia por documento), Z=
    Nondividend Distributions (casilla 3 del 1099, el ROC). Por eso el parser detecta la HOJA
    y la FILA de cabecera dinámicamente (buscando "Ticker Symbol"), pero usa columnas fijas.
  - 2023 (.pdf): reporte en dos "mitades" de páginas — las primeras N páginas traen el bloque
    izquierdo (Ticker, fechas, Total Distribution Per Share) y las últimas N páginas el bloque
    derecho (Box 2a..2f, con "Nondividend Distributions" = Box 3). Sin columnas de atribución
    de año (el documento entero es "2023 YEAR-END..."), así que toda fila atribuye a ese año.
    Las tablas de cada mitad alinean fila a fila por nombre de fondo (verificado: 0 mismatches
    en los 4 pares fila/página del documento real).

Uso:  ./.venv/bin/python tools/fetch_roc_ici.py [--out knowledge/roc_ici.yaml]
"""
import argparse
import datetime
import os
import re
import sys
import tempfile

import openpyxl
import pdfplumber
import requests
import yaml

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_PATH = os.path.join(HERE, "knowledge", "roc_ici.yaml")

BASE_URL = ("https://yieldmaxetfs.com/wp-content/uploads/TaxDocuments/"
            "All%20Funds%20Tax%20Documents/")
INDEX_URL = "https://yieldmaxetfs.com/tax-documents/"

# Metadatos de los tres documentos. `url` son las verificadas el 2026-08-21 (ver traspaso);
# si alguna cambia/da 404, la página índice (`INDEX_URL`) trae los enlaces bajo
# "All Funds Tax Documents" — no inventar una URL nueva.
ICI_DOCS = (
    {
        "year": 2023,
        "kind": "pdf",
        "filename": "2023 ICI Primary - YieldMax ETFs Final.pdf",
        "url": BASE_URL + "2023%20ICI%20Primary%20-%20YieldMax%20ETFs%20Final.pdf",
    },
    {
        "year": 2024,
        "kind": "xlsx",
        "filename": "2024 ICI Primary Layout - YieldMax.xlsx",
        "url": BASE_URL + "2024%20ICI%20Primary%20Layout%20-%20YieldMax.xlsx",
    },
    {
        "year": 2025,
        "kind": "xlsx",
        "filename": "2025 ICI Primary Layout – YieldMax.xlsx",
        "url": BASE_URL + "2025%20ICI%20Primary%20Layout%20%E2%80%93%20YieldMax.xlsx",
    },
)

_BROWSER_HEADERS = {
    "User-Agent": ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
                   "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"),
    "Accept": "*/*",
    "Referer": "https://yieldmaxetfs.com/tax-documents/",
}

# Columnas fijas (1-indexadas, como las da openpyxl) — medidas en 2024 y confirmadas
# idénticas en 2025. Ver docstring del módulo.
COL_TICKER = 3
COL_EXDIV = 8
COL_TOTAL = 10
COL_YEAR_PRIOR = 11   # K
COL_YEAR_NEXT = 12    # L
COL_YEAR_CURRENT = 13  # M
COL_ROC = 26          # Z — "Nondividend Distributions"

# Columnas del PDF de 2023 (0-indexadas, como las da pdfplumber `extract_tables()`).
PDF_COL_TICKER = 2
PDF_COL_TOTAL = 6
PDF_COL_ROC = 5   # dentro de la tabla del bloque DERECHO (Box 2a..2f)


def download_all(dest_dir):
    """Descarga los tres documentos a `dest_dir`. Devuelve {year: local_path}."""
    paths = {}
    for doc in ICI_DOCS:
        dest = os.path.join(dest_dir, doc["filename"])
        resp = requests.get(doc["url"], headers=_BROWSER_HEADERS, timeout=60)
        resp.raise_for_status()
        with open(dest, "wb") as fh:
            fh.write(resp.content)
        paths[doc["year"]] = dest
        print(f"{doc['year']}: descargado {doc['filename']} ({len(resp.content)} bytes)",
              file=sys.stderr)
    return paths


def _num(v):
    return float(v) if isinstance(v, (int, float)) else 0.0


def _find_header_row(ws, max_scan_rows=30):
    """Fila donde la columna `COL_TICKER` dice literalmente "Ticker Symbol". No asume un
    número de fila fijo: en 2024 es la 13, en 2025 la 12."""
    for r in range(1, max_scan_rows + 1):
        v = ws.cell(row=r, column=COL_TICKER).value
        if isinstance(v, str) and v.strip() == "Ticker Symbol":
            return r
    return None


def _find_sheet(wb):
    """Hoja que trae la cabecera "Ticker Symbol". No asume el nombre: en 2024 es
    "Primary Layout FINAL"; en 2025 el archivo trae además una hoja "Cover" sin datos y la
    hoja de datos se llama "YM - ICI Primary Layout"."""
    for ws in wb.worksheets:
        header_row = _find_header_row(ws)
        if header_row is not None:
            return ws, header_row
    return None, None


def _year_columns(ws, header_row):
    """{columna: año} para K/L/M, leyendo el año real de la cabecera de cada documento
    (cambia cada año: en el de 2024, K="2023 (Prior Year)"; en el de 2025, K="2024 (Prior
    Year)"). Devuelve solo las columnas cuya cabecera trae un año reconocible."""
    cols = {}
    for c in (COL_YEAR_PRIOR, COL_YEAR_NEXT, COL_YEAR_CURRENT):
        header = ws.cell(row=header_row, column=c).value or ""
        m = re.search(r"(\d{4})", str(header))
        if m:
            cols[c] = int(m.group(1))
    return cols


def parse_xlsx(path):
    """[(ticker, año, total_j, roc_z), ...] desde un ICI Primary Layout en Excel.

    El año de cada fila NO es el de `Ex-Dividend Date`: son las columnas K/L/M ("Year
    Included in Shareholders' Income") las que declaran a qué año fiscal se atribuye el
    importe de esa fila — una distribución puede pagarse en enero y reclasificarse al año
    anterior. Se agrupa por la columna que trae el importe, con fallback a `Ex-Dividend Date`
    solo si ninguna de las tres trae valor (no debería pasar; ver test_roc_ici.py)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws, header_row = _find_sheet(wb)
    if ws is None:
        raise ValueError(f"{path}: no se encontró ninguna hoja con cabecera 'Ticker Symbol'")
    year_cols = _year_columns(ws, header_row)

    rows_out = []
    for r in range(header_row + 1, ws.max_row + 1):
        tk = ws.cell(row=r, column=COL_TICKER).value
        if not isinstance(tk, str) or not tk.strip().isalpha():
            continue
        tk = tk.strip().upper()
        total = ws.cell(row=r, column=COL_TOTAL).value
        if not isinstance(total, (int, float)):
            continue
        roc = _num(ws.cell(row=r, column=COL_ROC).value)

        year = None
        for c, y in year_cols.items():
            v = ws.cell(row=r, column=c).value
            if isinstance(v, (int, float)) and v:
                year = y
                break
        if year is None:
            exdiv = ws.cell(row=r, column=COL_EXDIV).value
            year = getattr(exdiv, "year", None)
        if year is None:
            continue

        rows_out.append((tk, year, float(total), roc))
    return rows_out


def parse_pdf(path, doc_year):
    """[(ticker, año, total_j, roc_z), ...] desde un ICI Primary Layout en PDF (formato 2023).

    El documento reparte cada fila en dos "mitades" de páginas: las primeras N páginas traen
    el bloque izquierdo (Ticker + Total Distribution Per Share) y las últimas N el bloque
    derecho (Box 2a..2f, con "Nondividend Distributions" = Box 3, el ROC). Se emparejan por
    índice de fila DENTRO de cada par de páginas, verificando que el nombre de fondo en columna
    0 coincide en ambas mitades — si no coincide en algún par, es que el layout cambió y no se
    puede aparear con confianza: se lanza un error explícito en vez de arriesgar mezclar ROC
    de un fondo con el total de otro.

    El documento no trae columnas de atribución de año (es "AÑO YEAR-END TAX REPORTING
    INFORMATION" completo): toda fila atribuye a `doc_year`."""
    with pdfplumber.open(path) as pdf:
        n = len(pdf.pages)
        if n % 2 != 0:
            raise ValueError(f"{path}: {n} páginas (impar) — el layout esperado es mitad "
                              "bloque izquierdo + mitad bloque derecho (medido en el PDF de "
                              "2023). No se puede aparear con confianza.")
        title = pdf.pages[0].extract_text() or ""
        m = re.search(r"(\d{4})\s+YEAR-END TAX REPORTING INFORMATION", title)
        if m and int(m.group(1)) != doc_year:
            raise ValueError(f"{path}: el título del documento dice {m.group(1)} pero se "
                              f"esperaba {doc_year} — revisa que la URL/año no se cruzaron.")

        half = n // 2
        rows_out = []
        for i in range(half):
            left_tables = pdf.pages[i].extract_tables()
            right_tables = pdf.pages[i + half].extract_tables()
            if not left_tables or not right_tables:
                raise ValueError(f"{path}: página {i} o {i + half} sin tabla detectada por "
                                  "pdfplumber (¿cambió el layout del PDF?).")
            tl, tr = left_tables[0], right_tables[0]
            names_l = [row[0] for row in tl]
            names_r = [row[0] for row in tr]
            if names_l != names_r:
                raise ValueError(f"{path}: las tablas de las páginas {i}/{i + half} no "
                                  "alinean fila a fila por nombre de fondo — no se puede "
                                  "aparear ROC (bloque derecho) con su fila (bloque "
                                  "izquierdo) sin arriesgar mezclar fondos.")
            for rl, rr in zip(tl, tr):
                name = rl[0] or ""
                if not name or "Total" in name:
                    continue
                tk = (rl[PDF_COL_TICKER] or "").strip().upper()
                if not tk or not tk.isalpha():
                    continue
                total = _num_str(rl[PDF_COL_TOTAL])
                roc = _num_str(rr[PDF_COL_ROC])
                rows_out.append((tk, doc_year, total, roc))
    return rows_out


def _num_str(s):
    if s in (None, ""):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def aggregate(tagged_rows):
    """tagged_rows: [(ticker, año, j, z, source_doc, source_url), ...] ->
    {ticker: {año: {"j":.., "z":.., "docs": {...}, "urls": {...}}}}"""
    agg = {}
    for tk, year, j, z, doc, url in tagged_rows:
        bucket = agg.setdefault(tk, {}).setdefault(
            year, {"j": 0.0, "z": 0.0, "docs": set(), "urls": set()})
        bucket["j"] += j
        bucket["z"] += z
        bucket["docs"].add(doc)
        bucket["urls"].add(url)
    return agg


def build_yaml_entries(agg, asof):
    out = {}
    for tk, years in agg.items():
        out[tk] = {}
        for year, d in sorted(years.items()):
            if d["j"] <= 0:
                continue
            pct = round(d["z"] / d["j"] * 100, 2)
            docs = sorted(d["docs"])
            urls = sorted(d["urls"])
            out[tk][year] = {
                "roc_pct": pct,
                "source_url": urls[0] if len(urls) == 1 else urls,
                "source_doc": docs[0] if len(docs) == 1 else docs,
                "asof": asof,
            }
    return out


def main(argv):
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", default=OUT_PATH)
    args = parser.parse_args(argv[1:])

    with tempfile.TemporaryDirectory(prefix="roc_ici_") as tmp:
        paths = download_all(tmp)
        tagged = []
        for doc in ICI_DOCS:
            path = paths[doc["year"]]
            if doc["kind"] == "xlsx":
                rows = parse_xlsx(path)
            else:
                rows = parse_pdf(path, doc["year"])
            for tk, year, j, z in rows:
                tagged.append((tk, year, j, z, doc["filename"], doc["url"]))
            print(f"{doc['year']}: {len(rows)} filas parseadas de {doc['filename']}",
                  file=sys.stderr)

    agg = aggregate(tagged)
    asof = datetime.date.today().isoformat()
    out = build_yaml_entries(agg, asof)

    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as fh:
        fh.write("# Generado por tools/fetch_roc_ici.py (fuente: ICI Primary Layout de "
                  "YieldMax, https://yieldmaxetfs.com/tax-documents/).\n"
                  "# ROC definitivo (Box 3 del 1099-DIV / Nondividend Distributions), por "
                  "ticker y AÑO FISCAL. No editar a mano.\n"
                  "# NO está cableado a ninguna vista — ver "
                  "Obsidian/IA/traspaso-2026-08-21-roc-historico-ici.md.\n")
        yaml.safe_dump(out, fh, sort_keys=False, allow_unicode=True)
    print(f"Escrito {args.out} ({len(out)} fondos).", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
